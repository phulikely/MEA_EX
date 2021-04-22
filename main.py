import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_binary
import pytest
from io import BytesIO
from PIL import Image, ImageDraw
import pandas as pd
import settings
import openpyxl


# Execution
class TestMySite():

	@pytest.fixture()
	def set_up(self):
		self.driver = webdriver.Chrome('./chromedriver')
		self.driver.implicitly_wait(10)
		self.driver.maximize_window()
		self.driver.delete_all_cookies()
		yield
		self.driver.close()
		self.driver.quit()

	def screenshot(self):
		return self.driver.get_screenshot_as_png()

	def get_file_name_by_time(self, path):
		self.timestr = time.strftime("%Y%m%d%H%M%S")
		return path + 'screenshot_' + self.timestr + '.png'

	def draw_rectangle(self, element):
		location = element.location
		size = element.size
		png = self.screenshot()
		img = Image.open(BytesIO(png))
		left = location['x']
		top = location['y']
		right = location['x'] + size['width']
		bottom = location['y'] + size['height']
		red_frame = ImageDraw.Draw(img)
		red_frame = red_frame.rectangle((left - settings.OFFSET_ELEMENT, 
											top - settings.OFFSET_ELEMENT, 
											right + settings.OFFSET_ELEMENT, 
											bottom + settings.OFFSET_ELEMENT), 
											outline ="red", width=4)
		# img.save(file_name)
		return img


	def test_TC_LOGIN(self, set_up):
		wbk_name = 'test_data.xlsx'
		wbk = openpyxl.load_workbook(wbk_name)
		ws1 = wbk['Login']
		column_B = ws1['B']
		# all_xlsx = pd.read_excel('test_data.xlsx', sheet_name='Login')
		# all_xlsx2 = pd.read_excel('test_data.xlsx', sheet_name='Register')
		for row in range(2, len(column_B) + 1):
			expected = 'Invalid email or password!'
			try:
				if ws1.cell(row, 4).value == 'YES':
					path = self.get_file_name_by_time(settings.LOGIN_PATH)
					self.driver.get(settings.MAIN_URL)
					start_btn = self.driver.find_element_by_id('start')
					time.sleep(2)
					self.draw_rectangle(start_btn).save(path)
					time.sleep(2)
					start_btn.click()
					self.driver.find_element_by_name('email').clear()
					email = self.driver.find_element_by_name('email')
					email.send_keys(ws1.cell(row, 2).value)
					path = self.get_file_name_by_time(settings.LOGIN_PATH)
					self.draw_rectangle(email).save(path)
					time.sleep(2)
					self.driver.find_element_by_name('password').clear()
					password = self.driver.find_element_by_name('password')
					password.send_keys(ws1.cell(row, 3).value)
					path = self.get_file_name_by_time(settings.LOGIN_PATH)
					self.draw_rectangle(password).save(path)
					time.sleep(2)
					login_btn = self.driver.find_element_by_id('login')
					path = self.get_file_name_by_time(settings.LOGIN_PATH)
					self.draw_rectangle(login_btn).save(path)
					time.sleep(2)
					login_btn.click()
					WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, "Add Project"))).click()
					ws1.cell(row, 5).value = 'PASS'
					
			except:
				ws1.cell(row, 5).value = 'FAIL'
				err_msg = self.driver.find_element_by_xpath('//label')
				path = self.get_file_name_by_time(settings.LOGIN_PATH)
				self.draw_rectangle(err_msg).save(path)
				time.sleep(2)
				assert err_msg.text == expected		

		wbk.save(wbk_name)
		wbk.close
